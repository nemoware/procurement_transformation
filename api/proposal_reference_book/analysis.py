import base64
import logging
from io import BytesIO

from flask import abort
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet

from db.config import db_handle
from db.entity.lot import Lot
from db.entity.simple_entity import Segment, Sub_segment, Service, Unit, Rate, Stage

logger = logging.getLogger(__name__)


def compare_proposal(proposal_file, procurement_id):
    if procurement_id == '0':
        raise ValueError('Не используйте 0 ИД закупки')
    wb = load_workbook(filename=BytesIO(base64.b64decode(proposal_file)))
    try:
        end_index_row = excel_validation(wb)
    except ValueError as e:
        raise ValueError(str(e))
    ws: Worksheet = wb['Форма КП (для анализа рынка) ВР']
    subject = ws['C6'].value
    segment_name = ws['C8'].value
    sub_segment_name = ws['C9'].value
    service_code = ws['C10'].value
    service_name = ws['C11'].value

    subject: (object, bool) = ({'name': subject}, True)
    segment = Segment.get_or_create(name=segment_name)
    sub_segment = Sub_segment.get_or_create(name=sub_segment_name)
    service = Service.get_or_create(name=service_name)

    current_lots = (Lot.select(Lot, Segment, Sub_segment, Service, Stage, Rate, Unit)
                    .join(Segment).switch(Lot)
                    .join(Sub_segment).switch(Lot)
                    .join(Service).switch(Lot)
                    .join(Stage).switch(Lot)
                    .join(Rate).switch(Lot)
                    .join(Unit)
                    .where((Segment.name == segment_name) &
                           (Sub_segment.name == sub_segment_name) &
                           (Service.code == service_code)))

    list_of_lots_from_db = []

    for lot in current_lots:
        list_of_lots_from_db.append({
            'segment_name': lot.segment_id.name,
            'segment_id': lot.segment_id.id,
            'sub_segment_name': lot.sub_segment_id.name,
            'sub_segment_id': lot.sub_segment_id.id,
            'service_code': lot.service_code.code,
            'service_name': lot.service_code.name,
            'stage_name': lot.stage_id.name,
            'rate_name': lot.rate_id.name,
            'unit_name': lot.unit_id.name,
            'stage_id': lot.stage_id.id,
            'rate_id': lot.rate_id.id,
            'unit_id': lot.unit_id.id,
            'id': lot.procurement_id,
            'is_null': lot.is_null
        })
    list_of_lots_by_id = list(filter(lambda x: x['id'] == procurement_id, list_of_lots_from_db))

    del lot, current_lots

    list_of_lots_from_excel = []
    prev_stage_name = None
    for index, row in enumerate(ws.iter_rows(min_row=18, max_col=8, max_row=end_index_row - 1, values_only=True),
                                start=18):
        if prev_stage_name is None:
            prev_stage_name = row[1]
        if row[1] == 'Итого:':
            prev_stage_name = None
            continue
        list_of_lots_from_excel.append({
            'segment_name': segment_name,
            'segment_id': 0,
            'sub_segment_name': sub_segment_name,
            'sub_segment_id': 0,
            'service_code': service_code,
            'service_name': service_name,
            'stage_name': prev_stage_name,
            'rate_name': row[2],
            'unit_name': row[4],
            'stage_id': 0,
            'rate_id': 0,
            'unit_id': 0,
            'id': procurement_id,
            'is_null': False
        })

    response_data = {
        'fields': [{
            'label': 'subject',
            'value': subject[0]['name'],
            'reference_book': not subject[1]
        }, {
            'label': 'segment',
            'value': segment[0].name,
            'reference_book': not segment[1]
        }, {
            'label': 'sub_segment',
            'value': sub_segment[0].name,
            'reference_book': not sub_segment[1]
        }, {
            'label': 'service_code',
            'value': service[0].code,
            'reference_book': not service[1]
        }, {
            'label': 'service_name',
            'value': service[0].name,
            'reference_book': not service[1]
        }],
        'stages': []
    }
    stages = []
    for index, lot in enumerate(list_of_lots_from_excel):
        stage = next((item for item in stages if item['name'] == lot['stage_name']), None)
        if stage is None:
            stage = {
                'name': lot['stage_name'],
                'reference_book': any(item for item in list_of_lots_from_db if item['stage_name'] == lot['stage_name']),
                'rows': []
            }
            stages.append(stage)
        # row = next((item for item in stage['rows'] if item['name'] == lot['']))
        if stage['reference_book'] is False:
            stage['rows'].append({
                'rate': lot['rate_name'],
                'rate_reference_book': False,
                'unit': lot['unit_name'],
                'unit_reference_book': False
            })
        else:
            stage['rows'].append({
                'rate': lot['rate_name'],
                'rate_reference_book': any(
                    item for item in list_of_lots_from_db
                    if item['stage_name'] == lot['stage_name'] and
                    item['rate_name'] == lot['rate_name']
                ),
                'unit': lot['unit_name'],
                'unit_reference_book': any(
                    item for item in list_of_lots_from_db
                    if item['stage_name'] == lot['stage_name'] and
                    item['rate_name'] == lot['rate_name'] and
                    item['unit_name'] == lot['unit_name']
                )
            })

    response_data['stages'] = stages

    list_of_lots_from_excel = list(filter(
        lambda x: next(
            (item for item in list_of_lots_from_db
             if item['stage_name'] == x['stage_name'] and
             item['rate_name'] == x['rate_name'] and
             item['unit_name'] == x['unit_name']),
            False
        ),
        list_of_lots_from_excel
    ))

    for lot in list_of_lots_from_excel:
        lot['segment_id'] = next(
            (lot2['segment_id'] for lot2 in list_of_lots_from_db if lot['segment_name'] == lot2['segment_name'])
        )
        lot['sub_segment_id'] = next(
            (lot2['sub_segment_id'] for lot2 in list_of_lots_from_db if
             lot['sub_segment_name'] == lot2['sub_segment_name'])
        )
        lot['stage_id'] = next(
            (lot2['stage_id'] for lot2 in list_of_lots_from_db if lot['stage_name'] == lot2['stage_name'])
        )
        lot['rate_id'] = next(
            (lot2['rate_id'] for lot2 in list_of_lots_from_db if lot['rate_name'] == lot2['rate_name'])
        )
        lot['unit_id'] = next(
            (lot2['unit_id'] for lot2 in list_of_lots_from_db if lot['unit_name'] == lot2['unit_name'])
        )
    with db_handle.atomic() as transaction:
        try:
            # TODO
            # dict_list_of_lots_by_id = {}
            # for lot in list_of_lots_by_id:
            #     dict_list_of_lots_by_id[lot['unit']] = {}
            list_of_lots_without_duplicate = list({v['id']: v for v in list_of_lots_by_id}.values())
            delete_query = Lot.delete().where(Lot.procurement_id == procurement_id)
            delete_query.execute()
            for lot in list_of_lots_from_excel:
                Lot.create(procurement_id=procurement_id,
                           segment_id=lot["segment_id"],
                           sub_segment_id=lot["sub_segment_id"],
                           service_code=lot["service_code"],
                           stage_id=lot['stage_id'],
                           rate_id=lot['rate_id'],
                           unit_id=lot['unit_id'])
        except Exception as e:
            transaction.rollback()
            logger.exception(e)
            abort(500, e)
    return response_data


def excel_validation(wb: Workbook):
    ws: Worksheet = wb['Форма КП (для анализа рынка) ВР']

    for num in [6, 8, 9, 10, 11]:
        if ws[f'C{num}'].value is None:
            raise ValueError(f'Ошибка в шапке файла, а именно в ячейке C{num}')
    words = [
        '№ п\п',
        'Наименование этапа / услуги (работы)',
        'Наименование расценки',
        'Альтернативное наименование расценки',
        'Единица измерения\n(ЕИ)',
        'Кол-во ЕИ',
        'Стоимость ЕИ\nв валюте Участника',
        'ИТОГО\nв валюте Участника',
    ]
    for ind, cell in enumerate(ws[15]):
        if cell.value != words[ind]:
            raise ValueError(f'Ошибка в наименовании столбца, должно быть {words[0]}, что было найдено {cell.value}')
    for ind, cell in enumerate(ws[17], start=1):
        if ind != cell.value:
            raise ValueError(f'В строке под номером 17, не правильно указаны числа')

    is_stopped = False
    prev_stage_name = None
    end_index_row = 0
    for index, row in enumerate(ws.iter_rows(min_row=18, max_col=8, values_only=True), start=18):
        if is_stopped:
            break
        for ind, cell in enumerate(row, start=1):
            if ind == 1 and (row[1] == ('Командировочные расходы \n'
                                        '(При необходимости. Включаются в стоимость КП, '
                                        'если в ТЗ не заявлено требование об оказании услуг/выполнения работ в '
                                        'удаленном формате и/или для оказании услуг/выполнения работ требуется '
                                        'оформление командировки для указанного выше состава исполнителей)') or
                             row[1] == 'Всего без НДС'):
                is_stopped = True
                end_index_row = index
                break

            if ind == 1:
                if cell != f'=ROW(A{index})-17':
                    raise ValueError(f'Не правильно указана формула в 1 столбец, в строке {index}, '
                                     f'должно быть значение =ROW(A{index})-17, а было найдено {cell}')
            if ind == 2:
                if cell == 'Итого:':
                    prev_stage_name = None
                    break
                if prev_stage_name is None:
                    if type(cell) is not str and len(cell) == 0:
                        raise ValueError(f'Не был указан этап в строке {index}')
                    prev_stage_name = cell

            if ind == 3 and (type(cell) is not str or cell is None or len(cell) == 0):
                raise ValueError(f'Не правильно указан этап в строке {index}')

            if ind == 5 and (type(cell) is not str or cell is None or len(cell) == 0):
                raise ValueError(f'Не правильно указана расценка в строке {index}')

            if (ind == 6 or ind == 7) and ((type(cell) is not int and type(cell) is not float) or cell < 0):
                raise ValueError(
                    f'Не правильно указано значение в {"Кол-во ЕИ" if ind == 6 else "Стоимость ЕИ в валюте Участника"},'
                    f'в строке {index}')

            if ind == 8 and cell != f'=F{index}*G{index}':
                raise ValueError(
                    f'Ошибка в формуле в столбец \"ИТОГО в валюте Участника\", '
                    f'в строке {index}')
    return end_index_row
