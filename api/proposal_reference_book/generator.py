import base64
import logging
import re
from copy import copy
from datetime import datetime

from flask import abort
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.writer.excel import save_virtual_workbook
from peewee import fn, SQL

from api.common import env_var
from db.entity.lot import Lot
from db.entity.simple_entity import Segment, Sub_segment, Service, Unit, Rate, Stage

max_number_of_stages = env_var('MAX_NUMBER_OF_STAGES', 10)
max_number_of_rates = env_var('MAX_NUMBER_OF_RATES', 20)
logger = logging.getLogger(__name__)


def generate_prefilled_proposal(segment_name=None, sub_segment_name=None, service_code=None, service_name=None,
                                subject=None, guaranteed_volume=None):
    list_of_stage_ids = []

    try:
        service_code = int(service_code)
    except Exception as ex:
        msg = 'Не верно указан код услуги'
        logger.error(msg)
        abort(400, msg)

    current_stage_ids = (Lot.select(Lot.stage_id, fn.COUNT(Lot.stage_id).alias('num_stage_id'))
                         .join(Segment).switch(Lot)
                         .join(Sub_segment).switch(Lot)
                         .join(Service).switch(Lot)
                         .where(Segment.name == segment_name, Sub_segment.name == sub_segment_name,
                                Service.code == service_code)
                         .group_by(Lot.stage_id)
                         .order_by(SQL('num_stage_id').desc()))

    for lot in current_stage_ids:
        list_of_stage_ids.append({
            'stage_id': lot.stage_id_id,
            'num_stage_id': lot.num_stage_id
        })

    list_of_stage_ids = list_of_stage_ids[:max_number_of_stages]

    current_rate_ids = (
        Lot.select(Lot.rate_id, Lot.stage_id, Lot.is_null,
                   fn.COUNT(Lot.stage_id).alias('num_rate_id'))
        .join(Segment).switch(Lot)
        .join(Sub_segment).switch(Lot)
        .join(Service).switch(Lot)
        .where((Segment.name == segment_name) &
               (Sub_segment.name == sub_segment_name) &
               (Service.code == service_code) &
               (Lot.stage_id << list(map(lambda x: x['stage_id'], list_of_stage_ids))))
        .group_by(Lot.rate_id, Lot.stage_id, Lot.is_null)
        .order_by(SQL('num_rate_id').desc()))

    dict_of_rate_ids = {}
    for lot in current_rate_ids:
        dict_of_rate_ids.setdefault(lot.stage_id_id, []).append({
            'rate_id': lot.rate_id_id,
            'num_rate_id': lot.num_rate_id,
            'is_null': lot.is_null
        })

    list_of_rate_ids = []
    for stage in dict_of_rate_ids:
        count = 0
        for rate in dict_of_rate_ids[stage]:
            if count == max_number_of_rates:
                break
            else:
                list_of_rate_ids.append(rate['rate_id'])
                if rate['is_null']:
                    continue
                count += 1

    list_of_rate_ids = list(set(list_of_rate_ids))

    current_lots = (Lot.select(Lot,
                               Segment.name,
                               Sub_segment.name,
                               Service.code,
                               Stage.name,
                               Stage.id,
                               Rate.name,
                               Rate.id,
                               Unit.name,
                               Unit.id)
                    .join(Segment).switch(Lot)
                    .join(Sub_segment).switch(Lot)
                    .join(Service).switch(Lot)
                    .join(Stage).switch(Lot)
                    .join(Rate).switch(Lot)
                    .join(Unit)
                    .where((Segment.name == segment_name) &
                           (Sub_segment.name == sub_segment_name) &
                           (Service.code == service_code) &
                           (Lot.stage_id << list(map(lambda x: x['stage_id'], list_of_stage_ids))) &
                           (Lot.rate_id << list(map(lambda x: x, list_of_rate_ids))))
                    )

    list_of_lots = []
    if len(list_of_stage_ids) == 0 or len(list_of_rate_ids) == 0:
        msg = 'По данному запросу не были найдены этапы или расценки'
        logger.error(msg)
        abort(400, msg)

    del list_of_stage_ids, list_of_rate_ids, current_rate_ids, current_stage_ids

    for lot in current_lots:
        list_of_lots.append({
            'id': lot.procurement_id,
            'segment_name': lot.segment_id.name,
            'sub_segment_name': lot.sub_segment_id.name,
            'service_code': lot.service_code.code,
            'stage_name': lot.stage_id.name,
            'stage_id': lot.stage_id.id,
            'rate_name': lot.rate_id.name,
            'rate_id': lot.rate_id.id,
            'unit_name': lot.unit_id.name,
            'unit_id': lot.unit_id.id,
        })
    del current_lots

    current_lots = (Lot.select(Lot,
                               Segment.name,
                               Sub_segment.name,
                               Service,
                               Stage,
                               Rate,
                               Unit)
                    .join(Segment).switch(Lot)
                    .join(Sub_segment).switch(Lot)
                    .join(Service).switch(Lot)
                    .join(Stage).switch(Lot)
                    .join(Rate).switch(Lot)
                    .join(Unit)
                    .where((Segment.name == segment_name) &
                           (Sub_segment.name == sub_segment_name) &
                           (Service.code == service_code)))

    list_of_all_lots = []
    for lot in current_lots:
        list_of_all_lots.append({
            'segment_name': lot.segment_id.name,
            'sub_segment_name': lot.sub_segment_id.name,
            'service_code': lot.service_code.code,
            'service_name': lot.service_code.name,
            'number_of_service': 0,
            'stage_name': lot.stage_id.name,
            'number_of_stages': 0,
            'percent_of_stages': 0,
            'rate_name': lot.rate_id.name,
            'number_of_rates': 0,
            'percent_of_rates': 0,
            'unit_name': lot.unit_id.name,
            'number_of_units': 0,
            'percent_of_units': 0,
            'stage_id': lot.stage_id.id,
            'rate_id': lot.rate_id.id,
            'unit_id': lot.unit_id.id,
            'id': lot.procurement_id,
            'is_null': lot.is_null
        })
    del current_lots

    number_of_all_stages = {}
    list_if_ids = list(set(map(lambda x: x['id'], list_of_all_lots)))
    number_of_ids = len(list_if_ids) if '0' not in list_if_ids else len(list_if_ids) - 1
    list_of_stage_name = []

    for id in list_if_ids:
        lots = [single_lot for single_lot in list_of_all_lots if single_lot['id'] == id]
        for lot in lots:
            increase_number = 1
            if lot['is_null']:
                increase_number = 0

            stage_name = lot['stage_name']
            rate_name = lot['rate_name']
            unit_name = lot['unit_name']
            number_of_all_stages.setdefault(stage_name, {
                'count': 0
            })
            if stage_name not in list_of_stage_name:
                number_of_all_stages[stage_name]['count'] += increase_number
                list_of_stage_name.append(stage_name)

            number_of_all_stages[stage_name].setdefault(rate_name, {
                'count': 0
            })
            number_of_all_stages[stage_name][rate_name]['count'] = len(set([
                single_lot['id'] for single_lot in list_of_all_lots
                if single_lot['stage_name'] == stage_name and
                   single_lot['rate_name'] == rate_name and
                   single_lot['is_null'] is None
            ]))
            number_of_all_stages[stage_name][rate_name].setdefault(unit_name, {
                'count': 0
            })['count'] += increase_number
        list_of_stage_name = []

    del list_if_ids, lots, stage_name, rate_name, unit_name, id

    for lot in list_of_all_lots:
        lot['number_of_service'] = number_of_ids
        lot['number_of_stages'] = number_of_all_stages[lot['stage_name']]['count']
        lot['number_of_rates'] = number_of_all_stages[lot['stage_name']][lot['rate_name']]['count']
        if lot['is_null']:
            lot['number_of_units'] = 0
        else:
            lot['number_of_units'] = number_of_all_stages[lot['stage_name']][lot['rate_name']][lot['unit_name']][
                'count']
        lot['percent_of_stages'] = lot['number_of_stages'] / (number_of_ids / 100)
        lot['percent_of_rates'] = lot['number_of_rates'] / (lot['number_of_stages'] / 100)
        lot['percent_of_units'] = lot['number_of_units'] / (lot['number_of_rates'] / 100)

    list_of_lots = list(filter(lambda x: x['id'] != '0', list_of_lots))
    list_of_lots.sort(
        key=lambda x: (
            number_of_all_stages[x['stage_name']]['count'],
            number_of_all_stages[x['stage_name']][x['rate_name']]['count'],
            number_of_all_stages[x['stage_name']][x['rate_name']][x['unit_name']]['count']
        ), reverse=True
    )
    list_without_duplicates_of_lots = []
    for element in list_of_lots:
        del element['id']
    del element

    seen = set()
    for d in list_of_lots:
        t = tuple(d.items())
        if t not in seen:
            seen.add(t)
            list_without_duplicates_of_lots.append(d)
    for lot in list_without_duplicates_of_lots:
        del lot['rate_id'], lot['stage_id'], lot['unit_id']
    del seen, d, t, list_of_lots, lot

    list_of_all_lots.sort(
        key=lambda x: (
            number_of_all_stages[x['stage_name']]['count'],
            number_of_all_stages[x['stage_name']][x['rate_name']]['count'],
            number_of_all_stages[x['stage_name']][x['rate_name']][x['unit_name']]['count'] if
            not x['is_null'] else 0
        ), reverse=True
    )

    for element in list_of_all_lots:
        del element['id'], element['stage_id'], element['rate_id'], element['unit_id'], element['is_null'],
    del element, number_of_ids

    list_without_duplicates_of_all_lots = []
    seen = set()
    for d in list_of_all_lots:
        t = tuple(d.items())
        if t not in seen:
            seen.add(t)
            list_without_duplicates_of_all_lots.append(d)
    del seen, d, t, list_of_all_lots

    list_without_duplicates_of_lots.sort(key=lambda x: x['stage_name'])
    list_without_duplicates_of_lots.sort(key=lambda x: (
        number_of_all_stages[x['stage_name']]['count']
    ), reverse=True)

    list_without_duplicates_of_lots = list(filter(
        lambda x:
        x['unit_name'] == find_max_number_of_units(x, number_of_all_stages),
        list_without_duplicates_of_lots))

    this_is_base64 = generate_sheets(list_without_duplicates_of_lots,
                                     list_without_duplicates_of_all_lots,
                                     subject,
                                     segment_name,
                                     sub_segment_name,
                                     service_code,
                                     service_name)

    return {
        'proposal_file': this_is_base64,
        'name': f"КП {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}.xlsm",
        'size': int((len(this_is_base64) * 3) / 4)
    }


def get_excel_by_procurement_id(procurement_id):
    current_lots = (Lot.select(Lot, Segment.name, Sub_segment.name, Service, Stage, Rate, Unit)
                    .join(Segment).switch(Lot)
                    .join(Sub_segment).switch(Lot)
                    .join(Service).switch(Lot)
                    .join(Stage).switch(Lot)
                    .join(Rate).switch(Lot)
                    .join(Unit)
                    .where((Lot.procurement_id == procurement_id)))
    list_of_lots = []
    service_name = None
    for lot in current_lots:
        if service_name is None:
            service_name = lot.service_code.name
        list_of_lots.append({
            'segment_name': lot.segment_id.name,
            'sub_segment_name': lot.sub_segment_id.name,
            'service_code': lot.service_code.code,
            'stage_name': lot.stage_id.name,
            'rate_name': lot.rate_id.name,
            'unit_name': lot.unit_id.name,
        })
    if len(list_of_lots) == 0:
        raise ValueError('Такого ИД закупки не существует')
    list_without_duplicates_of_lots = []
    seen = set()
    for d in list_of_lots:
        t = tuple(d.items())
        if t not in seen:
            seen.add(t)
            list_without_duplicates_of_lots.append(d)

    this_is_base64 = generate_sheets(list_without_duplicates_of_lots,
                                     [],
                                     '>=D',
                                     list_of_lots[0]['segment_name'],
                                     list_of_lots[0]['sub_segment_name'],
                                     list_of_lots[0]['service_code'],
                                     service_name)

    return {
        'proposal_file': this_is_base64,
        'name': f"{str(procurement_id).upper()} {datetime.now().strftime('%d %H:%M:%S')}.xlsm",
        'size': int((len(this_is_base64) * 3) / 4)
    }


def is_duplicate(lot, lot2) -> bool:
    return (lot2['is_null'] is not True and
            lot['is_null'] is not True and
            lot2['segment_name'] == lot['segment_name'] and
            lot2['sub_segment_name'] == lot['sub_segment_name'] and
            lot2['service_code'] == lot['service_code'])


def copy_style(cell, new_cell):
    if cell.has_style:
        new_cell.font = copy(cell.font)
        new_cell.border = copy(cell.border)
        new_cell.fill = copy(cell.fill)
        new_cell.number_format = copy(cell.number_format)
        new_cell.protection = copy(cell.protection)
        new_cell.alignment = copy(cell.alignment)


def find_max_number_of_units(lot, number_of_all_stages):
    current_unit = ''
    max_number_of_units = 0
    for unit in number_of_all_stages[lot['stage_name']][lot['rate_name']]:
        if unit != 'count':
            if int(number_of_all_stages[lot['stage_name']][lot['rate_name']][unit]['count']) > max_number_of_units:
                current_unit = unit
                max_number_of_units = int(number_of_all_stages[lot['stage_name']][lot['rate_name']][unit]['count'])
    return current_unit


def generate_sheets(list_without_duplicates_of_lots, list_without_duplicates_of_all_lots, subject=None,
                    segment_name=None, sub_segment_name=None, service_code=None, service_name=None):
    workbook = load_workbook(filename=f"Пустой шаблон.xlsm", read_only=False,
                             keep_vba=True)
    ws: Worksheet = workbook["Форма КП (для анализа рынка) ВР"]
    index_of_number = 1
    ws['C6'] = subject
    ws.row_dimensions[6].height = 15 * get_height_for_row2(subject, max_length=125)
    ws['C8'] = segment_name
    ws.row_dimensions[8].height = 15 * get_height_for_row2(segment_name, max_length=125)
    ws['C9'] = sub_segment_name
    ws.row_dimensions[9].height = 15 * get_height_for_row2(sub_segment_name, max_length=125)
    ws['C10'] = service_code
    ws['C11'] = service_name
    ws.row_dimensions[11].height = 15 * get_height_for_row2(service_name, max_length=125)

    prev_stage_name = None
    start_index = 0
    end_index = 0
    shift_index = 0
    total_height_of_one_segment = 0
    for index, lot in enumerate(list_without_duplicates_of_lots, start=18):
        if prev_stage_name is None:
            prev_stage_name = lot['stage_name']
            start_index = index
        if prev_stage_name == lot['stage_name']:
            end_index = index + shift_index
        else:
            ws.merge_cells(start_row=start_index, start_column=2, end_row=end_index, end_column=2)
            ws.append([
                f'=ROW(A{index + shift_index})-17',
                'Итого:', '', '', '',
                f'=SUM(F{start_index}:F{end_index})',
                f'=SUM(G{start_index}:G{end_index})',
                f'=SUM(H{start_index}:H{end_index})',
            ])
            set_final_line_by_segment(end_index, ws)

            total_height_of_one_segment = 0
            for row in range(start_index, end_index + 1):
                total_height_of_one_segment += ws.row_dimensions[row].height

            row_height = total_height_of_one_segment - (
                    get_height_for_row2(ws[f'B{start_index}'].value, max_length=40) * 15)
            if row_height < 0:
                ws.row_dimensions[start_index].height = 62.25 + (
                        get_height_for_row2(ws[f'B{start_index}'].value, max_length=40) * 15)

            shift_index += 1
            start_index = index + shift_index
            end_index = index + shift_index
            prev_stage_name = lot['stage_name']
            index_of_number += 1

        lot = list(lot.values())
        del lot[0:3]
        lot.insert(0, f'=ROW(A{index + shift_index})-17')
        lot.insert(3, 'Заполняется при необходимости. См. пояснения п.3.')
        lot.extend([0, 0, f'=F{index + shift_index}*G{index + shift_index}'])
        ws.append(lot)

        coefficient_for_height: int = 0
        for ind, cell in enumerate(ws[end_index]):
            cell.font = Font(name='Arial', size=10)
            if ind == 1:
                cell.fill = PatternFill(fgColor="B7DEE8", fill_type="solid")
            if ind in [2, 3]:
                if ind == 2:
                    coefficient_for_height = get_height_for_row2(cell.value, max_length=31)
                if ind == 3:
                    cell.font = Font(name='Arial', size=10, color="595959", italic=True)
                cell.alignment = Alignment(vertical='center', wrapText=True)
            else:
                cell.alignment = Alignment(vertical='center',
                                           horizontal='center',
                                           wrapText=True)
            if ind in [5, 6, 7]:
                cell.number_format = numbers.FORMAT_NUMBER_00
            cell.border = Border(left=Side(style='medium'),
                                 right=Side(style='medium'),
                                 top=Side(style='medium'),
                                 bottom=Side(style='medium'))
            row_height = (15 * coefficient_for_height)
            ws.row_dimensions[end_index].height = 62.25 if 62.25 - row_height > 0 else row_height
        index_of_number += 1

    ws.merge_cells(start_row=start_index, start_column=2, end_row=end_index, end_column=2)
    ws.append([
        f'=ROW(A{end_index + 1})-17',
        'Итого:', '', '', '',
        f'=SUM(F{start_index}:F{end_index})',
        f'=SUM(G{start_index}:G{end_index})',
        f'=SUM(H{start_index}:H{end_index})',
    ])
    set_final_line_by_segment(end_index, ws)
    index_of_number += 1
    ws2 = workbook['footer']

    if any(item for item in list_without_duplicates_of_lots if
           item['unit_name'] in ['Ч/Д', 'Ч/Ч', 'Ч/М', 'Ч/С', 'ЧЕЛ']):
        end_index += 1
        for column_index in range(1, 9):
            cell = ws2.cell(row=19, column=column_index)
            new_cell = ws.cell(row=end_index + 1, column=column_index)
            if column_index == 1:
                new_cell.value = f'=ROW(A{end_index + 1})-17'
            else:
                new_cell.value = copy(cell.value)
            copy_style(cell, new_cell)
        ws.merge_cells(start_row=end_index + 1,
                       start_column=2,
                       end_row=end_index + 1,
                       end_column=3)
        ws.row_dimensions[end_index + 1].height = 68.25

    for row_index in range(1, 16):
        for column_index in range(1, 9):
            cell = ws2.cell(row=row_index, column=column_index)
            new_cell = ws.cell(row=row_index + 1 + end_index, column=column_index)
            new_cell.value = copy(cell.value)
            copy_style(cell, new_cell)
    workbook.remove_sheet(ws2)

    ws[f'F{end_index + 4}'].value = f'=F{end_index + 2}'
    ws[f'H{end_index + 4}'].value = f'=H{end_index + 2}+H{end_index + 3}'
    for ind in range(2, 5):
        ws.merge_cells(start_row=end_index + ind,
                       start_column=2,
                       end_row=end_index + ind,
                       end_column=3)
        ws.row_dimensions[end_index + ind].height = 30

    ws.merge_cells(start_row=end_index + 6,
                   start_column=1,
                   end_row=end_index + 7,
                   end_column=8)
    ws.row_dimensions[end_index + 6].height = 72
    ws.row_dimensions[end_index + 7].height = 70.50

    for ind in [10, 13, 16]:
        ws.merge_cells(start_row=end_index + ind,
                       start_column=3,
                       end_row=end_index + ind,
                       end_column=5)
        ws.row_dimensions[end_index + ind].height = 21.75

    ws = workbook["Справочник"]
    for index, lot in enumerate(list_without_duplicates_of_all_lots, start=2):
        ws.append(list(lot.values()))
        for ind, cell in enumerate(ws[index]):
            if ind in [4, 6, 7, 9, 10, 12, 13]:
                cell.alignment = Alignment(horizontal='right')
        if index % 2 == 1:
            for ind, cell in enumerate(ws[index]):
                cell.fill = PatternFill(fgColor="FDE9D9", fill_type="solid")
    workbook.active = workbook['Форма КП (для анализа рынка) ВР']
    virtual_workbook = save_virtual_workbook(workbook)
    this_is_base64 = base64.b64encode(virtual_workbook).decode('UTF-8')

    return this_is_base64


def set_final_line_by_segment(end_index, ws):
    ws.merge_cells(start_row=end_index + 1, start_column=2, end_row=end_index + 1, end_column=4)
    ws.row_dimensions[end_index + 1].height = 23.25
    for ind, cell in enumerate(ws[end_index + 1]):
        cell.border = Border(left=Side(style='medium'),
                             right=Side(style='medium'),
                             top=Side(style='medium'),
                             bottom=Side(style='medium'))
        cell.alignment = Alignment(vertical='center')
        if ind in [5, 6, 7]:
            cell.number_format = numbers.FORMAT_NUMBER_00
            cell.alignment = Alignment(vertical='center', horizontal='center')
        if ind == 0:
            cell.alignment = Alignment(vertical='center', horizontal='center')


def get_height_for_row2(value: str, max_length: int):
    re_compile = re.compile(r'([^a-zA-ZА-Яа-я\(\)]+|[a-zA-ZА-Яа-я\(\)])')
    only_special_charter = re.compile(r'([^a-zA-ZА-Яа-я\(\)]+)')
    split = re_compile.split(value)
    split: list[str] = list(filter(lambda x: x != '', split))
    number_of_lines = 1
    acc = ''
    for letter in split:
        if len(acc) < max_length:
            acc += letter
        else:
            number_of_lines += 1
            charter_split = only_special_charter.split(acc)
            charter_split.insert(0, letter)
            charter_split = list(filter(lambda x: x != '', charter_split))
            if charter_split:
                last_special_charter = charter_split[-1]
                acc = last_special_charter

    return number_of_lines
